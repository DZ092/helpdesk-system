"""Cobertura de fumaça: toda rota responde, e o ciclo do chamado atravessa o sistema.

A suíte de `test_app.py` verifica regras — quem pode o quê, o que é recusado.
Estes testes cobrem o outro lado: se cada tela ainda *carrega*. É o tipo de
falha que uma refatoração introduz e nenhum teste de regra percebe, porque o
template quebra sem que nenhuma regra mude.
"""

import io

import pytest
from openpyxl import load_workbook

SENHA = "senha-de-teste"

# Rotas que qualquer visitante alcança.
ROTAS_PUBLICAS = ["/login", "/cadastro", "/chamado", "/esqueci-senha"]

# Rotas que exigem sessão. O perfil mínimo de cada uma está anotado ao lado.
ROTAS_INTERNAS = [
    "/dashboard",
    "/chamados",
    "/meus-chamados",
    "/senha",
    "/meu-token",
    "/admin/usuarios",
    "/admin/logs",
]

# Combinações de filtro e paginação da listagem — o ponto onde um `url_for`
# desatualizado ou um parâmetro inesperado costuma derrubar a página.
CONSULTAS = [
    "?pagina=1",
    "?pagina=999",
    "?status=Aberto",
    "?prioridade=Crítica",
    "?setor=Financeiro",
    "?responsavel=nenhum",
    "?responsavel=abacaxi",
    "?busca=impressora",
    "?data_inicio=2020-01-01&data_fim=2030-12-31",
]


def entrar(client, email, senha=SENHA):
    return client.post("/login", data={"email": email, "senha": senha})


def entrar_como_admin(client, criar_usuario):
    criar_usuario(nome="Ana Admin", email="admin@teste.com", tipo="Administrador")
    entrar(client, "admin@teste.com")


def abrir_chamado(client, titulo="Impressora não imprime"):
    return client.post(
        "/chamado",
        data={
            "usuario": "Rafael Souza",
            "setor": "Financeiro",
            "titulo": titulo,
            "descricao": "A impressora aceita o trabalho mas nada sai.",
            "prioridade": "Alta",
        },
    )


@pytest.mark.parametrize("rota", ROTAS_PUBLICAS)
def test_pagina_publica_responde(client, rota):
    assert client.get(rota).status_code == 200


def test_raiz_redireciona_para_o_dashboard(client):
    resposta = client.get("/")
    assert resposta.status_code == 302
    assert resposta.headers["Location"].endswith("/dashboard")


@pytest.mark.parametrize("rota", ROTAS_INTERNAS)
def test_rota_interna_exige_sessao(client, rota):
    """Sem login, toda tela interna devolve para o login."""
    resposta = client.get(rota, follow_redirects=True)
    assert b"Login" in resposta.data


@pytest.mark.parametrize("rota", ROTAS_INTERNAS)
def test_rota_interna_responde_para_admin(client, criar_usuario, rota):
    entrar_como_admin(client, criar_usuario)
    assert client.get(rota).status_code == 200


@pytest.mark.parametrize("consulta", CONSULTAS)
def test_listagem_aguenta_filtros_e_paginacao(client, criar_usuario, consulta):
    entrar_como_admin(client, criar_usuario)
    abrir_chamado(client)
    assert client.get(f"/chamados{consulta}").status_code == 200


def test_detalhe_de_chamado_inexistente_devolve_404(client, criar_usuario):
    entrar_como_admin(client, criar_usuario)
    assert client.get("/chamados/9999").status_code == 404


def test_ciclo_do_chamado_atravessa_o_sistema(client, criar_usuario):
    """Abertura pública, atendimento, comentário e resolução — ponta a ponta."""
    entrar_como_admin(client, criar_usuario)
    abrir_chamado(client, titulo="Notebook não conecta na VPN")

    detalhe = client.get("/chamados/1")
    assert detalhe.status_code == 200
    assert "Notebook não conecta na VPN" in detalhe.get_data(as_text=True)

    client.post("/chamados/1/status", data={"status": "Em andamento"})
    client.post("/chamados/1/comentarios", data={"mensagem": "Certificado reinstalado."})
    client.post("/chamados/1/status", data={"status": "Resolvido"})

    depois = client.get("/chamados/1").get_data(as_text=True)
    assert "Certificado reinstalado." in depois
    assert "Resolvido" in depois

    # O chamado atendido passa a aparecer na tela do responsável.
    assert "Notebook" in client.get("/meus-chamados").get_data(as_text=True)


def test_logs_registram_o_que_aconteceu(client, criar_usuario):
    """A trilha de auditoria é a única tela que nenhum outro teste visita."""
    entrar_como_admin(client, criar_usuario)
    abrir_chamado(client)
    client.post("/chamados/1/status", data={"status": "Em andamento"})

    logs = client.get("/admin/logs").get_data(as_text=True)
    assert "Login realizado" in logs
    assert "Abertura de chamado" in logs


def test_telas_de_autenticacao_carregam_os_dois_css(client):
    """O auth.css vale só no login e no cadastro; o style.css, em tudo."""
    for rota in ("/login", "/cadastro"):
        html = client.get(rota).get_data(as_text=True)
        assert "css/style.css" in html
        assert "css/auth.css" in html


def test_telas_de_dados_nao_carregam_o_auth_css(client, criar_usuario):
    entrar_como_admin(client, criar_usuario)
    html = client.get("/dashboard").get_data(as_text=True)
    assert "css/style.css" in html
    assert "css/auth.css" not in html


# ==============================================================================
# EXPORTAÇÃO DE RELATÓRIOS (issue #7)
# ==============================================================================
def test_exportar_exige_login(client):
    resposta = client.get("/chamados/exportar?formato=excel", follow_redirects=True)
    assert b"Login" in resposta.data


def test_exportar_recusa_usuario_comum(client, criar_usuario):
    """Mais restrita que a própria listagem: exige perfil Técnico/Administrador."""
    criar_usuario()
    entrar(client, "fulano@teste.com")

    resposta = client.get("/chamados/exportar?formato=excel", follow_redirects=True)
    assert b"restrita a T" in resposta.data


def test_exportar_excel_responde_para_tecnico(client, criar_usuario):
    criar_usuario(nome="Ana Técnica", email="tecnica@teste.com", tipo="Técnico")
    entrar(client, "tecnica@teste.com")
    abrir_chamado(client)

    resposta = client.get("/chamados/exportar?formato=excel")

    assert resposta.status_code == 200
    assert resposta.mimetype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in resposta.headers["Content-Disposition"]


def test_exportar_pdf_responde_para_admin(client, criar_usuario):
    entrar_como_admin(client, criar_usuario)
    abrir_chamado(client)

    resposta = client.get("/chamados/exportar?formato=pdf")

    assert resposta.status_code == 200
    assert resposta.mimetype == "application/pdf"


@pytest.mark.parametrize("consulta", CONSULTAS)
def test_exportar_aguenta_os_mesmos_filtros_da_listagem(client, criar_usuario, consulta):
    """Os mesmos parâmetros que a listagem aceita não podem derrubar o export."""
    entrar_como_admin(client, criar_usuario)
    abrir_chamado(client)

    resposta = client.get(f"/chamados/exportar{consulta}&formato=excel")
    assert resposta.status_code == 200


def test_exportar_recusa_formato_desconhecido(client, criar_usuario):
    entrar_como_admin(client, criar_usuario)

    resposta = client.get("/chamados/exportar?formato=csv", follow_redirects=True)
    assert resposta.status_code == 200
    assert "formato de exporta" in resposta.get_data(as_text=True).lower()


def test_exportacao_registra_log(client, criar_usuario):
    entrar_como_admin(client, criar_usuario)
    abrir_chamado(client)

    client.get("/chamados/exportar?formato=pdf")

    logs = client.get("/admin/logs").get_data(as_text=True)
    assert "Exportação de relatório" in logs


def test_exportacao_respeita_o_teto_de_linhas(client, criar_usuario, monkeypatch):
    """Sem o teto, o mesmo filtro vazio que a tela pagina em 15 por vez viraria
    uma exportação sem limite nenhum — aqui reduzimos o teto para não precisar
    criar milhares de chamados só para exercitar a regra."""
    import rotas.chamados as modulo_chamados

    monkeypatch.setattr(modulo_chamados, "LIMITE_EXPORTACAO", 3)

    entrar_como_admin(client, criar_usuario)
    for i in range(6):
        abrir_chamado(client, titulo=f"Chamado {i}")

    resposta = client.get("/chamados/exportar?formato=excel")
    pasta = load_workbook(io.BytesIO(resposta.data))
    linhas = list(pasta.active.iter_rows(values_only=True))

    assert len(linhas) - 1 == 3  # cabeçalho + 3 linhas, não as 6 criadas


# ==============================================================================
# UPLOAD DE ANEXOS (issue #6)
# ==============================================================================
def _imagem_falsa(nome="foto.png"):
    return (io.BytesIO(b"conteudo-de-imagem-falso"), nome)


def _mockar_upload(monkeypatch, url="https://res.cloudinary.com/teste/imagem.png"):
    """Substitui a chamada real ao Cloudinary por uma URL fixa.

    `enviar_anexo` é importado por nome dentro de `rotas.chamados`
    (`from armazenamento import enviar_anexo`), então o monkeypatch precisa
    mirar a referência já importada ali — corrigir só `armazenamento.enviar_anexo`
    não afetaria a rota, que já guarda o próprio ponteiro para a função.
    """
    import rotas.chamados as modulo_chamados

    monkeypatch.setattr(modulo_chamados, "enviar_anexo", lambda arquivo: url)


def test_abertura_de_chamado_aceita_anexo(client, monkeypatch):
    _mockar_upload(monkeypatch)

    resposta = client.post(
        "/chamado",
        data={
            "usuario": "Rafael Souza",
            "setor": "Financeiro",
            "titulo": "Impressora não imprime",
            "descricao": "A impressora aceita o trabalho mas nada sai.",
            "prioridade": "Alta",
            "anexos": [_imagem_falsa()],
        },
        content_type="multipart/form-data",
    )
    # A abertura bem-sucedida agora renderiza a tela de confirmação com o
    # código de acompanhamento (ver rotas/chamados.py::chamado), em vez de
    # redirecionar de volta para /chamado.
    assert resposta.status_code == 200

    from extensions import db
    from models import Anexo

    anexos = db.session.execute(db.select(Anexo)).scalars().all()
    assert len(anexos) == 1
    assert anexos[0].comentario_id is None
    assert anexos[0].nome_original == "foto.png"


def test_abertura_de_chamado_sem_anexo_continua_funcionando(client, monkeypatch):
    """O campo é opcional — nada deve chamar o Cloudinary se nenhum arquivo vier."""
    chamado = False

    def _falha_se_chamado(arquivo):
        nonlocal chamado
        chamado = True
        return "não deveria ter sido chamado"

    import rotas.chamados as modulo_chamados

    monkeypatch.setattr(modulo_chamados, "enviar_anexo", _falha_se_chamado)

    resposta = abrir_chamado(client)
    assert resposta.status_code == 200
    assert chamado is False


def test_abertura_de_chamado_ignora_extensao_nao_permitida(client, monkeypatch):
    _mockar_upload(monkeypatch)

    client.post(
        "/chamado",
        data={
            "usuario": "Rafael Souza",
            "setor": "Financeiro",
            "titulo": "Chamado com anexo inválido",
            "descricao": "Descrição qualquer.",
            "prioridade": "Alta",
            "anexos": [_imagem_falsa(nome="virus.exe")],
        },
        content_type="multipart/form-data",
    )

    from extensions import db
    from models import Anexo

    assert db.session.execute(db.select(Anexo)).scalars().all() == []


def test_abertura_de_chamado_respeita_o_teto_de_anexos(client, monkeypatch):
    """Seis arquivos enviados de uma vez — só os 5 primeiros viram Anexo."""
    _mockar_upload(monkeypatch)

    client.post(
        "/chamado",
        data={
            "usuario": "Rafael Souza",
            "setor": "Financeiro",
            "titulo": "Chamado com vários anexos",
            "descricao": "Descrição qualquer.",
            "prioridade": "Alta",
            "anexos": [_imagem_falsa(nome=f"foto{i}.png") for i in range(6)],
        },
        content_type="multipart/form-data",
    )

    from extensions import db
    from models import Anexo

    assert len(db.session.execute(db.select(Anexo)).scalars().all()) == 5


def test_comentario_aceita_anexo(client, criar_usuario, monkeypatch):
    _mockar_upload(monkeypatch)
    entrar_como_admin(client, criar_usuario)
    abrir_chamado(client)

    client.post(
        "/chamados/1/comentarios",
        data={"mensagem": "Segue print do erro.", "anexos": [_imagem_falsa(nome="erro.jpg")]},
        content_type="multipart/form-data",
    )

    from extensions import db
    from models import Anexo, Comentario

    comentario = db.session.execute(db.select(Comentario)).scalars().one()
    anexo = db.session.execute(db.select(Anexo)).scalars().one()
    assert anexo.comentario_id == comentario.id
    assert anexo.chamado_id == 1


def test_detalhe_do_chamado_exibe_anexo_da_abertura(client, criar_usuario, monkeypatch):
    _mockar_upload(monkeypatch, url="https://res.cloudinary.com/teste/abertura.png")
    entrar_como_admin(client, criar_usuario)

    client.post(
        "/chamado",
        data={
            "usuario": "Rafael Souza",
            "setor": "Financeiro",
            "titulo": "Chamado com anexo visível",
            "descricao": "Descrição qualquer.",
            "prioridade": "Alta",
            "anexos": [_imagem_falsa()],
        },
        content_type="multipart/form-data",
    )

    html = client.get("/chamados/1").get_data(as_text=True)
    assert "https://res.cloudinary.com/teste/abertura.png" in html


def test_detalhe_do_chamado_exibe_anexo_do_comentario(client, criar_usuario, monkeypatch):
    entrar_como_admin(client, criar_usuario)
    abrir_chamado(client)
    _mockar_upload(monkeypatch, url="https://res.cloudinary.com/teste/comentario.png")

    client.post(
        "/chamados/1/comentarios",
        data={"mensagem": "Segue print.", "anexos": [_imagem_falsa()]},
        content_type="multipart/form-data",
    )

    html = client.get("/chamados/1").get_data(as_text=True)
    assert "https://res.cloudinary.com/teste/comentario.png" in html
