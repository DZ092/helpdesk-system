"""Geração dos relatórios em Excel e PDF, isolada da rota (issue #7)."""

import io

from openpyxl import load_workbook

from extensions import db
from models import Chamado
from relatorios import COLUNAS, gerar_excel, gerar_pdf


def _criar_chamado(**extra):
    dados = dict(
        usuario="Maria",
        setor="Financeiro",
        titulo="Computador não liga",
        descricao="Tela preta ao ligar",
        status="Aberto",
        prioridade="Alta",
    )
    dados.update(extra)
    chamado = Chamado(**dados)
    db.session.add(chamado)
    db.session.commit()
    return chamado


def test_excel_com_lista_vazia_tem_so_o_cabecalho(app):
    with app.app_context():
        conteudo = gerar_excel([])

    pasta = load_workbook(io.BytesIO(conteudo))
    aba = pasta.active
    linhas = list(aba.iter_rows(values_only=True))

    assert linhas == [COLUNAS]


def test_excel_traz_uma_linha_por_chamado(app):
    with app.app_context():
        c1 = _criar_chamado(titulo="Chamado um")
        c2 = _criar_chamado(titulo="Chamado dois", status="Resolvido")
        conteudo = gerar_excel([c1, c2])

    pasta = load_workbook(io.BytesIO(conteudo))
    aba = pasta.active
    linhas = list(aba.iter_rows(values_only=True))

    assert len(linhas) == 3  # cabeçalho + 2 chamados
    titulos = {linha[1] for linha in linhas[1:]}
    assert titulos == {"Chamado um", "Chamado dois"}


def test_excel_mostra_travessao_quando_sem_responsavel(app):
    with app.app_context():
        chamado = _criar_chamado()
        conteudo = gerar_excel([chamado])

    pasta = load_workbook(io.BytesIO(conteudo))
    linhas = list(pasta.active.iter_rows(values_only=True))
    responsavel = linhas[1][COLUNAS.index("Responsável")]
    assert responsavel == "—"


def test_pdf_com_lista_vazia_nao_quebra(app):
    """Um filtro que não bate com nenhum chamado ainda precisa devolver um PDF
    válido, com só a linha de cabeçalho — não um erro 500."""
    with app.app_context():
        conteudo = gerar_pdf([])

    assert conteudo.startswith(b"%PDF")


def test_pdf_com_chamados_gera_documento_valido(app):
    with app.app_context():
        chamado = _criar_chamado(titulo="Impressora não imprime")
        conteudo = gerar_pdf([chamado])

    assert conteudo.startswith(b"%PDF")
    assert len(conteudo) > 500  # um PDF vazio de verdade não passaria disso
