"""Exportação da listagem de chamados em Excel e PDF (issue #7).

Gera o mesmo recorte que a tela `/chamados` está mostrando — os filtros vêm
de `rotas.chamados.query_chamados_filtrados`, não de uma segunda consulta.
Cada formato é uma função pura que recebe a lista de chamados já carregada e
devolve os bytes prontos para download; quem decide o nome do arquivo e o
`mimetype` é a rota.
"""

import io
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from constantes import FUSO_EXIBICAO

COLUNAS = ("ID", "Título", "Setor", "Status", "Prioridade", "Responsável", "Criado em")


def _linha(chamado):
    """Uma tupla por chamado, na ordem de `COLUNAS` — usada pelos dois formatos.

    Mesma conversão do filtro de template `data_local` (`app.py`): o banco
    grava UTC "naive", então marcar `tzinfo=utc` antes de converter é o que
    faz o horário virar Brasília de verdade, em vez de ficar UTC disfarçado.
    """
    responsavel = chamado.responsavel.nome if chamado.responsavel else "—"
    criado_em = chamado.criado_em
    if criado_em.tzinfo is None:
        criado_em = criado_em.replace(tzinfo=timezone.utc)
    criado_em_local = criado_em.astimezone(FUSO_EXIBICAO)
    return (
        chamado.id,
        chamado.titulo,
        chamado.setor,
        chamado.status,
        chamado.prioridade,
        responsavel,
        criado_em_local.strftime("%d/%m/%Y %H:%M"),
    )


def gerar_excel(chamados_lista):
    """Devolve os bytes de uma planilha .xlsx com uma linha por chamado."""
    pasta = Workbook()
    aba = pasta.active
    aba.title = "Chamados"

    aba.append(COLUNAS)
    for celula in aba[1]:
        celula.font = Font(bold=True)

    for chamado in chamados_lista:
        aba.append(_linha(chamado))

    # Largura automática, limitada para o título e a descrição não esticarem
    # a coluna até ficar ilegível numa tela comum.
    for indice, cabecalho in enumerate(COLUNAS, start=1):
        maior = max(
            [len(cabecalho)] + [len(str(linha[indice - 1])) for linha in aba.iter_rows(min_row=2, values_only=True)]
        )
        aba.column_dimensions[get_column_letter(indice)].width = min(maior + 2, 40)

    buffer = io.BytesIO()
    pasta.save(buffer)
    return buffer.getvalue()


def gerar_pdf(chamados_lista):
    """Devolve os bytes de um PDF com a mesma listagem, em página paisagem.

    Paisagem porque a tabela tem sete colunas — em retrato o título e a
    descrição ficariam espremidos demais para servir de relatório gerencial.
    """
    buffer = io.BytesIO()
    documento = SimpleDocTemplate(
        buffer,
        pagesize=landscape(A4),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph("Relatório de chamados — Help Desk System", estilos["Title"]),
        Paragraph(
            f"Gerado em {datetime.now(FUSO_EXIBICAO).strftime('%d/%m/%Y %H:%M')} "
            f"— {len(chamados_lista)} chamado(s)",
            estilos["Normal"],
        ),
        Spacer(1, 0.5 * cm),
    ]

    dados_tabela = [list(COLUNAS)] + [list(_linha(c)) for c in chamados_lista]
    tabela = Table(dados_tabela, repeatRows=1)
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2d2d44")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f2f2f2")]),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]
        )
    )
    elementos.append(tabela)

    documento.build(elementos)
    return buffer.getvalue()
